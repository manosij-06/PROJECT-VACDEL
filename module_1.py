import random
from openpyxl import load_workbook

def username_password(username,password):
    flag = 0
    book = load_workbook("database1.xlsx")
    sheet1 = book["Admin"]
    sheet2 = book["Centre"]
    for i in range(2,sheet1.max_row+1):
        if sheet1.cell(row = i,column = 1).value == username:
            if sheet1.cell(row = i,column = 2).value == password:
                flag = 1
    for i in range(2,sheet2.max_row+1):
        if sheet2.cell(row = i,column = 1).value == username:
            if sheet2.cell(row = i,column = 2).value == password:
                flag = 2

    return flag

def recaptcha():
    re_captcha = ["Es4Ky","yK4sE","rDf8B","B8fDr","Ugi8a","a8igU","GGJu8","8uJGG","Jkl4s","s4lkJ","Oplh5","5hlpo","Xc5gh","hg5cX","ZxlD2","2DlxZ","Uyrd3","3dryU","Mbvc1","1cvbM","Bvc2a","a2cvB","Lvc1g","g1cvL","Gh3kl","lk3hG","Fgil4","4ligF","Shu5k","k5uhS","Tylf7","7flyT","Fgb3j","j3bgF","Elp5r","r5plE","plJ9d","d9Jlp","Lvbn7","7nbvL","jlu9M","M9ulj","mBn5e","e5nBm","kL0bd","db0Lk","cGop5","5poGc","snmI4","4Imns"]

    print("\n\t\t\t\t\t\t-----------")
    print("\t\t\t\t\t\t|  {recaptcha}  |".format(recaptcha = re_captcha[random.randrange(0,len(re_captcha))]))
    print("\t\t\t\t\t\t-----------")

    rc = input("\n\t\t\t   Enter the recaptcha : ")

    if rc in re_captcha:
        return
    else:
        print("\n\t\t\t\t\t   Invalid Captcha")
        recaptcha()
