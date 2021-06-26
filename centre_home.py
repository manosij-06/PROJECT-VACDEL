from openpyxl import load_workbook

def choice_2(user):
    print("\n\t\t<> Enter (E) for consignment entry\t\t<> Enter (R) to request for vaccines\t\t<> Enter (L) to logout")
    c = input("\n\t\t=>")
    if c.lower() == 'e':
        consg_entry(user)
    elif c.lower() == 'r':
        req_vccn(user)
    elif c.lower() == 'l':
        print("\n\n\t\t\t\t\t\t\t<> You are logged out . Login again ...\n")
        return 1
    choice_2(user)


def consg_entry(user):
    book = load_workbook("centre_database.xlsx")
    temp = user[:-9].split("_")
    temp = " ".join(temp) + "_centre " + user[-1]
    sheet_user = book[temp]
    phase_n = input("\n\t\t\t\t\t<> Phase No. : ")
    recv_date = input("\n\t\t\t\t\t<> Enter the date recieved in format (dd-mm-yyyy) : ")
    recv_vcc = input("\n\t\t\t\t\t<> Enter the no. of vaccines recieved : ")
    recv_good = input("\n\t\t\t\t\t<> Condition (GOOD) : ")
    recv_bad = input("\n\t\t\t\t\t<> Condition (BAD) : ")
    #phase_n = input("\n\t\t\t\t\t<> Phase No. : ")
    for i in range(2,sheet_user.max_row+1):
        if sheet_user.cell(row = i,column = 1).value == phase_n.upper():
            #print("SUCCESS")
            sheet_user.cell(row = i,column = 4).value = recv_date
            sheet_user.cell(row = i,column = 5).value = recv_vcc
            sheet_user.cell(row = i,column = 6).value = recv_good
            sheet_user.cell(row = i,column = 7).value = recv_bad
            sheet_user.cell(row = i,column = 8).value = input("\n\t\t\t\t\t<> Remarks if any :")
    book.save("centre_database.xlsx")

def req_vccn(user):
    file = open("resources.txt","a")
    vcc_n = input("\n\t\t\t\t\t<> Enter the number of vaccines required : ")
    st_date = input("\n\t\t\t\t\t<> Enter the stipulated date : ")
    sen = "\n\t\tCentre ID : " + user + " \t\tVaccines required : " + vcc_n + "\t\tWithin : " + st_date
    file.write(sen)
    file.close()
