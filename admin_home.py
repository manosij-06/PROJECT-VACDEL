from openpyxl import load_workbook


def choice():
    print("\t\t<> Enter (D) to show database\t\t<> Enter (E) for consignment dispatch entry\t\t<> Enter (L) to logout")
    c = input("\n\t\t=>")
    if c.lower() == 'd':
        show()
    elif c.lower() == 'l':
        print("\n\n\t\t\t\t\t\t\t<> You are logged out . Login again ...\n")
        return 1
    elif c.lower() == 'e':
        dispatch_entry()
    else:
        print("\n\t\t\t\t\t\t-----------------------------")
        print("\t\t\t\t\t\t|  Invalid input . Try again ...  |")
        print("\t\t\t\t\t\t-----------------------------")
        choice()
    print("\n\n")
    choice()


def show():
    book = load_workbook("Database_1.xlsx")
    sheet = book["Howrah"]
    print("\n")
    print("\n\n\t\t\t\t\t\t\t<> DISTRICT DATA <>")
    for i in range(1,sheet.max_row+1):
        print("\n")
        for j in range(1,sheet.max_column+1):
            print(str(sheet.cell(row=i,column=j).value).center(20),end = " ")
    sheet_2 = book["Divisions"]
    print("\n")
    print("\n\n\t\t\t\t\t\t\t<> DIVISIONS DATA <>")
    for i in range(1,sheet_2.max_row+1):
        print("\n")
        for j in range(1,sheet_2.max_column+1):
            print(str(sheet_2.cell(row = i,column = j).value).center(20),end = " ")
    print("\n\n")
    choice_3 = input("\t\t\t\t\t<> Do you want division info : Y - yes | N - No : ")
    if choice_3.lower() == 'y':
        show_div()

def show_div():
    book = load_workbook("Database_1.xlsx")
    print("\n\n")
    division = input("\t\t\t\t\t<> Enter the division for more info or (X) to main menu : ")
    if division.lower() == 'x':
        return
    sheet_3 = book[division.lower()]
    print("\n\n")
    for i in range(1,sheet_3.max_row+1):
        print("\n")
        for j in range(1,sheet_3.max_column+1):
            print(str(sheet_3.cell(row = i,column = j).value).center(20),end = " ")
    print("\n\n")
    choice_4 = input("\t\t\t\t\t<> Do you want centre info : Y - yes | N - No : ")
    if choice_4.lower() == 'y':
        show_cen_info(division.lower())
    print("\n")
    con = input("\t\t\t\t\t<> Do you want to continue with division data : Y - yes | N - No : ")
    if con.lower() == 'y':
        show_div()


def show_cen_info(division_name):
    book = load_workbook("centre_database.xlsx")
    print("\n\n")
    cen_name = input("\t\t\t\t\t<> Enter the centre name inside this division for more info or (X) to main menu : ")
    if cen_name.lower() == 'x':
        return
    data_name = division_name + "_" + cen_name.lower()
    sheet_4 = book[data_name]
    print("\n\n")
    for i in range(1,sheet_4.max_row+1):
        print("\n")
        for j in range(1,sheet_4.max_column+1):
            print(str(sheet_4.cell(row = i,column = j).value).center(20),end = " ")
    print("\n\n")
    con = input("\t\t\t\t\t<> Do you want to continue with centre data : Y - yes | N - No : ")
    if con.lower() == 'y':
        show_cen_info(division_name)


def dispatch_entry():
    show_imp_data()
    print("\n\n")
    book = load_workbook("centre_database.xlsx")
    division_name = input("\t\t\t\t\t<> Enter the division name : ")
    print("\n")
    centre_name = input("\t\t\t\t\t<> Enter the centre name : ")
    file_name = division_name.lower() + "_" + centre_name.lower()
    sheet_5 = book[file_name]
    print("\n")
    date_time = input("\t\t\t\t\t<> Enter the date in format (dd-mm-yyyy) : ")
    n_vcc = input("\n\t\t\t\t\t<> Enter the no. of vaccines sent : ")
    sheet_5.cell(row = sheet_5.max_row+1,column = 1).value = "PHASE " + str(sheet_5.max_row)
    sheet_5.cell(row = sheet_5.max_row,column = 2).value = " " + date_time
    sheet_5.cell(row = sheet_5.max_row,column = 3).value = n_vcc
    book.save("centre_database.xlsx")


def show_imp_data():
    print("\n\t\t\t\t\t<> SOME IMPORTANT INFO TO LOOK AT <>")
    print("\n")
    file = open("resources.txt","r")
    for line in file:
        print(line.center(20))
    file.close()
    seen = input("\n\t\t\t\t<> Enter (S) to mark it as seen : ")
    if seen.lower() == 's':
        file = open("resources.txt","w")
        file.close()
